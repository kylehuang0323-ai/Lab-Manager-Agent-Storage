"""
资产管理模块 — 固定资产全生命周期管理 (Excel 数据层)
参考 SAP RealTimeFixedAssetDetailReport 设计，精简为实用字段
"""

import os
import re
import threading
from datetime import datetime
from dateutil.relativedelta import relativedelta
from typing import Optional

from openpyxl import Workbook, load_workbook

import config

_lock = threading.Lock()

# --------------------------------------------------
# 资产表头
# --------------------------------------------------

ASSET_HEADERS = [
    "asset_id",          # 资产编号 (即 SAP Asset Tag，由用户提交)
    "serial_number",     # 序列号
    "name",              # 资产名称
    "description",       # 详细描述
    "category",          # 分类: Desktop/Surface/Monitor/Peripheral/AV/Furniture/Other
    "model",             # 型号
    "status",            # 状态: 在用/闲置/维修/借出/报废
    "custodian",         # 保管人
    "assigned_to",       # 当前使用人
    "department",        # 部门
    "building",          # 楼宇
    "room",              # 房间
    "location_detail",   # 具体位置描述
    "cost_basis",        # 采购成本
    "currency",          # 币种
    "po_number",         # PO 订单号
    "purchase_date",     # 采购/入资日期 (Asset Cap Date)
    "start_date",        # 启用日期 (= Asset Cap Date)
    "dispose_date",      # 报废日期 (= Asset Cap Date + MAP Aging)
    "asset_age",         # 资产年龄 (来自 SAP Asset Age)
    "map_aging",         # MAP 折旧周期 (来自 SAP MAP Aging)
    "useful_life",       # 使用年限
    "last_inventory_date",  # 最近盘点日期
    "processor",         # 处理器 (IT 资产)
    "memory",            # 内存 (IT 资产)
    "storage",           # 存储 (IT 资产)
    "notes",             # 备注
    "created_at",
    "updated_at",
]

# 资产操作流水
ASSET_TX_HEADERS = [
    "tx_id",             # 流水号 ATX-0001
    "asset_id",          # 资产编号
    "asset_name",        # 资产名称
    "type",              # assign/return/transfer/repair/dispose/inventory
    "from_person",       # 来源人
    "to_person",         # 目标人
    "operator",          # 操作人
    "note",              # 备注
    "timestamp",
]

ASSET_FILE = os.path.join(config.DATA_DIR, "assets.xlsx")
ASSET_TX_FILE = os.path.join(config.DATA_DIR, "asset_transactions.xlsx")


# --------------------------------------------------
# 通用 Excel 工具 (复用 inventory_manager 模式)
# --------------------------------------------------

def _ensure_workbook(filepath: str, headers: list) -> None:
    if os.path.exists(filepath):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = max(len(h) + 4, 14)
    wb.save(filepath)


def _read_all(filepath: str, headers: list) -> list[dict]:
    if not os.path.exists(filepath):
        return []
    with _lock:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append(dict(zip(headers, row)))
        wb.close()
    return rows


def _append_row(filepath: str, headers: list, data: dict) -> None:
    with _lock:
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append([data.get(h, "") for h in headers])
        wb.save(filepath)


def _update_row(filepath: str, headers: list, key_col: str, key_val, updates: dict) -> bool:
    key_idx = headers.index(key_col) + 1
    with _lock:
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[key_idx - 1].value) == str(key_val):
                for col_name, val in updates.items():
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        ws.cell(row=row[0].row, column=col_idx, value=val)
                wb.save(filepath)
                return True
        wb.save(filepath)
    return False


def _next_id(filepath: str, headers: list, id_col: str, prefix: str) -> str:
    rows = _read_all(filepath, headers)
    max_num = 0
    for r in rows:
        val = str(r.get(id_col, ""))
        if val.startswith(prefix + "-"):
            try:
                num = int(val.split("-")[1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass
    return f"{prefix}-{max_num + 1:04d}"


def init_asset_files():
    _ensure_workbook(ASSET_FILE, ASSET_HEADERS)
    _ensure_workbook(ASSET_TX_FILE, ASSET_TX_HEADERS)


# --------------------------------------------------
# 资产 CRUD
# --------------------------------------------------

def get_all_assets() -> list[dict]:
    return _read_all(ASSET_FILE, ASSET_HEADERS)


def get_asset(asset_id: str) -> Optional[dict]:
    for a in get_all_assets():
        if str(a.get("asset_id")) == asset_id:
            return a
    return None


def search_assets(keyword: str) -> list[dict]:
    keyword = keyword.lower()
    results = []
    for a in get_all_assets():
        searchable = " ".join([
            str(a.get(f, "")) for f in
            ["asset_id", "name", "description", "category", "model", "serial_number",
             "custodian", "assigned_to", "location_detail", "status"]
        ]).lower()
        if keyword in searchable:
            results.append(a)
    return results


def get_asset_categories() -> list[str]:
    cats = set()
    for a in get_all_assets():
        cat = a.get("category", "")
        if cat:
            cats.add(str(cat))
    return sorted(cats)


def get_assets_by_status(status: str) -> list[dict]:
    return [a for a in get_all_assets() if str(a.get("status", "")).lower() == status.lower()]


def create_asset(asset_id: str, name: str, category: str = "Other", **kwargs) -> dict:
    """新建资产，asset_id 由用户提交（即内部资产编号 / Asset Tag）"""
    init_asset_files()
    if not asset_id or not str(asset_id).strip():
        raise ValueError("必须提供资产编号 (asset_id)")
    asset_id = str(asset_id).strip()

    # 检查是否已存在
    if get_asset(asset_id):
        raise ValueError(f"资产编号 {asset_id} 已存在")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    asset = {
        "asset_id": asset_id,
        "name": name,
        "category": category,
        "status": kwargs.get("status", "闲置"),
        "created_at": now,
        "updated_at": now,
    }
    for field in ASSET_HEADERS:
        if field in kwargs and field not in asset:
            asset[field] = kwargs[field]
    _append_row(ASSET_FILE, ASSET_HEADERS, asset)
    return asset


def update_asset(asset_id: str, **updates) -> dict:
    """更新资产信息"""
    init_asset_files()
    asset = get_asset(asset_id)
    if not asset:
        raise ValueError(f"资产 {asset_id} 不存在")
    updates["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _update_row(ASSET_FILE, ASSET_HEADERS, "asset_id", asset_id, updates)
    asset.update(updates)
    return asset


# --------------------------------------------------
# 资产操作
# --------------------------------------------------

def assign_asset(asset_id: str, to_person: str, operator: str = "系统", note: str = "") -> dict:
    """分配资产给用户"""
    init_asset_files()
    asset = get_asset(asset_id)
    if not asset:
        raise ValueError(f"资产 {asset_id} 不存在")
    if asset.get("status") == "报废":
        raise ValueError(f"资产 {asset_id} 已报废，无法分配")

    from_person = str(asset.get("assigned_to", "") or "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    updates = {"assigned_to": to_person, "status": "在用", "updated_at": now}
    if not asset.get("start_date"):
        updates["start_date"] = now.split(" ")[0]
    _update_row(ASSET_FILE, ASSET_HEADERS, "asset_id", asset_id, updates)

    tx_id = _next_id(ASSET_TX_FILE, ASSET_TX_HEADERS, "tx_id", "ATX")
    tx = {
        "tx_id": tx_id, "asset_id": asset_id,
        "asset_name": asset.get("name", ""),
        "type": "assign", "from_person": from_person,
        "to_person": to_person, "operator": operator,
        "note": note, "timestamp": now,
    }
    _append_row(ASSET_TX_FILE, ASSET_TX_HEADERS, tx)
    return {"tx_id": tx_id, "asset_id": asset_id, "name": asset.get("name"), "assigned_to": to_person}


def return_asset(asset_id: str, operator: str = "系统", note: str = "") -> dict:
    """归还资产"""
    init_asset_files()
    asset = get_asset(asset_id)
    if not asset:
        raise ValueError(f"资产 {asset_id} 不存在")

    from_person = str(asset.get("assigned_to", "") or "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _update_row(ASSET_FILE, ASSET_HEADERS, "asset_id", asset_id, {
        "assigned_to": "", "status": "闲置", "updated_at": now,
    })

    tx_id = _next_id(ASSET_TX_FILE, ASSET_TX_HEADERS, "tx_id", "ATX")
    tx = {
        "tx_id": tx_id, "asset_id": asset_id,
        "asset_name": asset.get("name", ""),
        "type": "return", "from_person": from_person,
        "to_person": "", "operator": operator,
        "note": note, "timestamp": now,
    }
    _append_row(ASSET_TX_FILE, ASSET_TX_HEADERS, tx)
    return {"tx_id": tx_id, "asset_id": asset_id, "name": asset.get("name"), "returned_by": from_person}


def transfer_asset(asset_id: str, to_person: str, operator: str = "系统", note: str = "") -> dict:
    """转移资产"""
    init_asset_files()
    asset = get_asset(asset_id)
    if not asset:
        raise ValueError(f"资产 {asset_id} 不存在")

    from_person = str(asset.get("assigned_to", "") or "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _update_row(ASSET_FILE, ASSET_HEADERS, "asset_id", asset_id, {
        "assigned_to": to_person, "status": "在用", "updated_at": now,
    })

    tx_id = _next_id(ASSET_TX_FILE, ASSET_TX_HEADERS, "tx_id", "ATX")
    tx = {
        "tx_id": tx_id, "asset_id": asset_id,
        "asset_name": asset.get("name", ""),
        "type": "transfer", "from_person": from_person,
        "to_person": to_person, "operator": operator,
        "note": note, "timestamp": now,
    }
    _append_row(ASSET_TX_FILE, ASSET_TX_HEADERS, tx)
    return {"tx_id": tx_id, "asset_id": asset_id, "name": asset.get("name"),
            "from": from_person, "to": to_person}


def dispose_asset(asset_id: str, operator: str = "系统", note: str = "") -> dict:
    """报废资产"""
    init_asset_files()
    asset = get_asset(asset_id)
    if not asset:
        raise ValueError(f"资产 {asset_id} 不存在")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _update_row(ASSET_FILE, ASSET_HEADERS, "asset_id", asset_id, {
        "status": "报废", "assigned_to": "", "dispose_date": now.split(" ")[0], "updated_at": now,
    })

    tx_id = _next_id(ASSET_TX_FILE, ASSET_TX_HEADERS, "tx_id", "ATX")
    tx = {
        "tx_id": tx_id, "asset_id": asset_id,
        "asset_name": asset.get("name", ""),
        "type": "dispose", "from_person": str(asset.get("assigned_to", "") or ""),
        "to_person": "", "operator": operator,
        "note": note, "timestamp": now,
    }
    _append_row(ASSET_TX_FILE, ASSET_TX_HEADERS, tx)
    return {"tx_id": tx_id, "asset_id": asset_id, "name": asset.get("name"), "status": "报废"}


def mark_inventory(asset_id: str, operator: str = "系统", note: str = "") -> dict:
    """盘点资产"""
    init_asset_files()
    asset = get_asset(asset_id)
    if not asset:
        raise ValueError(f"资产 {asset_id} 不存在")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _update_row(ASSET_FILE, ASSET_HEADERS, "asset_id", asset_id, {
        "last_inventory_date": now, "updated_at": now,
    })

    tx_id = _next_id(ASSET_TX_FILE, ASSET_TX_HEADERS, "tx_id", "ATX")
    tx = {
        "tx_id": tx_id, "asset_id": asset_id,
        "asset_name": asset.get("name", ""),
        "type": "inventory", "from_person": "", "to_person": "",
        "operator": operator, "note": note or "盘点确认", "timestamp": now,
    }
    _append_row(ASSET_TX_FILE, ASSET_TX_HEADERS, tx)
    return {"tx_id": tx_id, "asset_id": asset_id, "name": asset.get("name"), "inventory_date": now}


def get_asset_transactions(asset_id: str = None, tx_type: str = None,
                           limit: int = 50) -> list[dict]:
    rows = _read_all(ASSET_TX_FILE, ASSET_TX_HEADERS)
    if asset_id:
        rows = [r for r in rows if str(r.get("asset_id")) == asset_id]
    if tx_type:
        rows = [r for r in rows if r.get("type") == tx_type]
    rows.sort(key=lambda r: r.get("timestamp", ""), reverse=True)
    return rows[:limit]


def get_asset_summary() -> dict:
    """资产概览统计"""
    assets = get_all_assets()
    total = len(assets)
    by_status = {}
    by_category = {}
    for a in assets:
        s = str(a.get("status", "未知"))
        c = str(a.get("category", "未分类"))
        by_status[s] = by_status.get(s, 0) + 1
        by_category[c] = by_category.get(c, 0) + 1
    return {
        "total": total,
        "by_status": by_status,
        "by_category": by_category,
    }


# --------------------------------------------------
# SAP Excel 导入
# --------------------------------------------------

# SAP 列名 → 系统字段映射
SAP_COLUMN_MAP = {
    "Asset Tag": "asset_tag",
    "Serial Number": "serial_number",
    "Asset Description": "name",
    "Standard PO Description": "description",
    "Asset Class Description": "category",
    "Custodian": "custodian",
    "Bldg": "building",
    "Room": "room",
    "Location": "location_detail",
    "Cost Basis": "cost_basis",
    "CURR": "currency",
    "PO Number": "po_number",
    "Asset Cap Date": "purchase_date",
    "Asset Age": "asset_age",
    "MAP Aging": "map_aging",
    "Useful Life": "useful_life",
    "Last Inventory Date": "last_inventory_date",
    "Processor Speed": "processor",
    "Physical Memory": "memory",
    "HD Storage": "storage",
    "details": "notes",
    "Asset Main Num": "asset_tag_fallback",  # 备选 tag
}

# More AssetMgr sheet 映射
ASSETMGR_COLUMN_MAP = {
    "AssetTagNbr": "asset_tag",
    "SN": "serial_number",
    "Description": "name",
    "Category": "category",
    "Model": "model",
    "Owner": "custodian",
    "Building": "building",
    "Location": "location_detail",
    "PO": "po_number",
    "ShippingDate": "purchase_date",
}


def _parse_aging_to_months(aging_str: str) -> Optional[int]:
    """解析 MAP Aging / Useful Life 字符串为月数。
    支持格式: '15 years', '3yr', '5 Years, 0 Months', '60 months' 等
    """
    if not aging_str:
        return None
    s = str(aging_str).strip().lower()
    total_months = 0
    # 匹配年份
    m_y = re.search(r'(\d+)\s*year', s)
    if m_y:
        total_months += int(m_y.group(1)) * 12
    # 简写 yr
    if not m_y:
        m_yr = re.search(r'(\d+)\s*yr', s)
        if m_yr:
            total_months += int(m_yr.group(1)) * 12
    # 匹配月份
    m_m = re.search(r'(\d+)\s*month', s)
    if m_m:
        total_months += int(m_m.group(1))
    return total_months if total_months > 0 else None


def _calc_dispose_date(cap_date_str: str, aging_str: str) -> str:
    """根据 Asset Cap Date + MAP Aging 计算报废日期"""
    if not cap_date_str or not aging_str:
        return ""
    months = _parse_aging_to_months(aging_str)
    if not months:
        return ""
    try:
        cap_date = datetime.strptime(str(cap_date_str).strip()[:10], "%Y-%m-%d")
        end_date = cap_date + relativedelta(months=months)
        return end_date.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return ""


def import_sap_excel(filepath: str) -> dict:
    """
    导入 SAP 固资报表 Excel，支持多 Sheet 格式
    自动识别 RTFixedAssetDetail / More AssetMgr / All 格式
    """
    init_asset_files()

    if not os.path.exists(filepath):
        return {"success": 0, "skipped": 0, "errors": [f"文件不存在: {filepath}"]}

    try:
        wb = load_workbook(filepath, read_only=True)
    except Exception as e:
        return {"success": 0, "skipped": 0, "errors": [f"无法读取: {e}"]}

    imported = []
    skipped = 0
    errors = []
    existing_ids = {str(a.get("asset_id", "")).lower() for a in get_all_assets() if a.get("asset_id")}
    existing_sns = {str(a.get("serial_number", "")).lower() for a in get_all_assets() if a.get("serial_number")}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h else "" for h in rows[0]]

        # 选择列映射
        col_map_template = None
        if "Asset Tag" in headers or "Asset Description" in headers:
            col_map_template = SAP_COLUMN_MAP
        elif "AssetTagNbr" in headers:
            col_map_template = ASSETMGR_COLUMN_MAP
        elif "Asset Tag" not in headers and "AssetTagNbr" not in headers:
            # 尝试简单格式 (All sheet)
            simple_map = {"AssetTagNbr": "asset_tag", "SN": "serial_number",
                          "Description": "name", "Category": "category",
                          "Model": "model", "Location": "location_detail"}
            if any(h in headers for h in simple_map):
                col_map_template = simple_map
            else:
                continue

        # 构建 列索引→字段名 映射
        col_map = {}
        for i, h in enumerate(headers):
            if h in col_map_template:
                col_map[i] = col_map_template[h]

        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                data = {}
                for col_idx, field in col_map.items():
                    if col_idx < len(row) and row[col_idx] is not None:
                        data[field] = row[col_idx]

                name = str(data.get("name", "")).strip()
                if not name:
                    continue

                tag = str(data.get("asset_tag", data.get("asset_tag_fallback", ""))).strip()
                sn = str(data.get("serial_number", "")).strip()

                # Asset Tag 作为资产编号，必须存在
                if not tag or tag in ("N/A", "None", ""):
                    skipped += 1
                    continue

                # 去重 (按 asset_id)
                if tag.lower() in existing_ids:
                    skipped += 1
                    continue
                if sn and sn != "N/A" and sn.lower() in existing_sns:
                    skipped += 1
                    continue

                # 清理
                if sn in ("N/A", "None", ""):
                    sn = ""

                # 推断状态
                status = "在用"
                loc = str(data.get("location_detail", "")).lower()
                if "报废" in name or "dispose" in loc:
                    status = "报废"
                elif "仓库" in loc or "闲置" in loc:
                    status = "闲置"

                # 计算日期：start_date 来自 Asset Cap Date，dispose_date 自动推算
                cap_date = str(data.get("purchase_date", ""))
                aging_raw = str(data.get("map_aging", ""))
                dispose_date = _calc_dispose_date(cap_date, aging_raw)

                asset = create_asset(
                    asset_id=tag,
                    name=name[:100],
                    category=_normalize_category(str(data.get("category", "Other"))),
                    serial_number=sn,
                    description=str(data.get("description", ""))[:200],
                    model=str(data.get("model", "")),
                    status=status,
                    custodian=str(data.get("custodian", "")),
                    building=str(data.get("building", "")),
                    room=str(data.get("room", "")),
                    location_detail=str(data.get("location_detail", "")),
                    cost_basis=data.get("cost_basis", ""),
                    currency=str(data.get("currency", "CNY")),
                    po_number=str(data.get("po_number", "")),
                    purchase_date=cap_date,
                    start_date=cap_date,
                    dispose_date=dispose_date,
                    asset_age=str(data.get("asset_age", "")),
                    map_aging=aging_raw,
                    useful_life=str(data.get("useful_life", "")),
                    last_inventory_date=str(data.get("last_inventory_date", "")),
                    processor=str(data.get("processor", "")),
                    memory=str(data.get("memory", "")),
                    storage=str(data.get("storage", "")),
                    notes=str(data.get("notes", "")),
                )
                imported.append(asset)
                existing_ids.add(tag.lower())
                if sn:
                    existing_sns.add(sn.lower())

            except Exception as e:
                errors.append(f"Sheet '{sheet_name}' 第{row_idx}行: {e}")

    wb.close()
    return {"success": len(imported), "skipped": skipped, "errors": errors, "items": imported}


def _normalize_category(raw: str) -> str:
    """标准化分类名"""
    raw_lower = raw.lower()
    if any(k in raw_lower for k in ["surface", "surface hub"]):
        return "Surface"
    if any(k in raw_lower for k in ["desktop", "台式", "主机", "optiplex"]):
        return "Desktop"
    if any(k in raw_lower for k in ["monitor", "显示器", "显示屏"]):
        return "Monitor"
    if any(k in raw_lower for k in ["macbook", "imac", "mac"]):
        return "Mac"
    if any(k in raw_lower for k in ["audio", "visual", "相机", "摄像", "投影", "speaker", "epson", "sony"]):
        return "AV Equipment"
    if any(k in raw_lower for k in ["furn", "furniture", "桌", "椅", "柜"]):
        return "Furniture"
    if any(k in raw_lower for k in ["lab", "laser", "激光", "切割", "3d", "打印"]):
        return "Lab Equipment"
    if any(k in raw_lower for k in ["hololens", "vr", "ar", "mixed reality"]):
        return "MR/VR Device"
    if any(k in raw_lower for k in ["computer", "laptop", "笔记本"]):
        return "Laptop"
    return raw if raw else "Other"
