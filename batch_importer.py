"""
批量导入 — 解析上传的 Excel 文件，批量添加商品到库存
"""

import os
from datetime import datetime
from typing import Tuple

from openpyxl import load_workbook

import inventory_manager as im

# 支持的列名映射（兼容中英文列名）
COLUMN_ALIASES = {
    "name":      ["name", "名称", "商品名称", "商品名", "品名"],
    "category":  ["category", "分类", "类别", "品类"],
    "quantity":  ["quantity", "数量", "库存", "库存数量", "qty"],
    "unit":      ["unit", "单位"],
    "location":  ["location", "位置", "存放位置", "库位"],
    "min_stock": ["min_stock", "最低库存", "最低", "安全库存", "告警阈值"],
}


def _match_column(header: str) -> str | None:
    """将 Excel 列头匹配到标准字段名"""
    h = header.strip().lower()
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias.lower() == h:
                return field
    return None


def parse_excel(filepath: str) -> Tuple[list[dict], list[str]]:
    """
    解析 Excel 文件，返回 (items, errors)
    items: 可以导入的商品列表
    errors: 解析过程中的错误信息
    """
    if not os.path.exists(filepath):
        return [], [f"文件不存在: {filepath}"]

    try:
        wb = load_workbook(filepath, read_only=True)
    except Exception as e:
        return [], [f"无法读取 Excel 文件: {str(e)}"]

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], ["文件为空或只有表头"]

    # 解析表头
    headers = rows[0]
    col_map = {}  # column_index → field_name
    for i, h in enumerate(headers):
        if h is None:
            continue
        field = _match_column(str(h))
        if field:
            col_map[i] = field

    if "name" not in col_map.values():
        return [], [f"未找到「名称」列。检测到的列头: {[str(h) for h in headers if h]}"]

    # 解析数据行
    items = []
    errors = []
    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            item = {}
            for col_idx, field in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                item[field] = val

            name = item.get("name")
            if not name or str(name).strip() == "":
                continue  # 跳过空行

            items.append({
                "name": str(name).strip(),
                "category": str(item.get("category", "未分类")).strip(),
                "quantity": int(item.get("quantity") or 0),
                "unit": str(item.get("unit", "个")).strip(),
                "location": str(item.get("location", "")).strip(),
                "min_stock": int(item.get("min_stock") or 0),
            })
        except Exception as e:
            errors.append(f"第 {row_idx} 行解析错误: {str(e)}")

    return items, errors


def batch_import(filepath: str) -> dict:
    """
    批量导入商品

    Returns:
        {
            "success": int,     # 成功导入数
            "skipped": int,     # 跳过（重名）
            "errors": list,     # 错误列表
            "items": list,      # 成功导入的商品
        }
    """
    im.init_data_files()
    items, parse_errors = parse_excel(filepath)

    if not items and parse_errors:
        return {"success": 0, "skipped": 0, "errors": parse_errors, "items": []}

    # 获取现有商品名（用于去重）
    existing_names = {str(it.get("name", "")).lower() for it in im.get_all_items()}

    imported = []
    skipped = 0
    import_errors = list(parse_errors)

    for item_data in items:
        if item_data["name"].lower() in existing_names:
            skipped += 1
            continue
        try:
            result = im.create_item(**item_data)
            imported.append(result)
            existing_names.add(item_data["name"].lower())
        except Exception as e:
            import_errors.append(f"导入 \"{item_data['name']}\" 失败: {str(e)}")

    return {
        "success": len(imported),
        "skipped": skipped,
        "errors": import_errors,
        "items": imported,
    }


def generate_template() -> str:
    """生成导入模板 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    headers = ["名称", "分类", "数量", "单位", "存放位置", "最低库存"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 16

    # 示例数据
    examples = [
        ["Surface 鼠标", "配件", 100, "个", "A区柜1", 10],
        ["Type-C 线", "配件", 50, "根", "A区柜2", 20],
        ["笔记本支架", "办公用品", 30, "个", "B区柜1", 5],
    ]
    for row_data in examples:
        ws.append(row_data)

    template_path = os.path.join(os.path.dirname(__file__), "exports", "import_template.xlsx")
    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    wb.save(template_path)
    return template_path
