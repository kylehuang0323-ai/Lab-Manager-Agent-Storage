"""
Excel 数据层 — 库存表 + 出入库记录表的 CRUD 操作
使用 openpyxl 读写 Excel，threading.Lock 保证并发安全
"""

import os
import threading
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook

import config

_lock = threading.Lock()

# --------------------------------------------------
# 表头定义
# --------------------------------------------------

INVENTORY_HEADERS = [
    "item_id", "name", "category", "quantity", "unit",
    "location", "min_stock", "created_at", "updated_at",
]

TRANSACTION_HEADERS = [
    "tx_id", "item_id", "item_name", "type", "quantity",
    "balance_after", "operator", "recipient", "note", "timestamp",
]


# --------------------------------------------------
# 初始化 Excel 文件
# --------------------------------------------------

def _ensure_workbook(filepath: str, headers: list) -> None:
    """如果文件不存在则创建并写入表头；若已存在则同步缺失的列"""
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(headers)
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = max(len(h) + 4, 14)
        wb.save(filepath)
        return
    # 已存在时检查是否需要补列
    with _lock:
        wb = load_workbook(filepath)
        ws = wb.active
        existing = [cell.value for cell in ws[1]]
        changed = False
        for h in headers:
            if h not in existing:
                col = len(existing) + 1
                ws.cell(row=1, column=col, value=h)
                ws.column_dimensions[ws.cell(1, col).column_letter].width = max(len(h) + 4, 14)
                existing.append(h)
                changed = True
        if changed:
            wb.save(filepath)
        else:
            wb.close()


def init_data_files():
    """确保数据文件就绪"""
    _ensure_workbook(config.INVENTORY_FILE, INVENTORY_HEADERS)
    _ensure_workbook(config.TRANSACTIONS_FILE, TRANSACTION_HEADERS)


# --------------------------------------------------
# 通用读写
# --------------------------------------------------

def _read_all_rows(filepath: str, headers: list) -> list[dict]:
    """读取所有数据行，返回 dict 列表（按 Excel 实际表头匹配，兼容列变更）"""
    if not os.path.exists(filepath):
        return []
    with _lock:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = []
        # 读取实际表头，按表头名匹配而非固定位置
        file_headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            raw = dict(zip(file_headers, row))
            rows.append({h: raw.get(h) for h in headers})
        wb.close()
    return rows


def _append_row(filepath: str, headers: list, data: dict) -> None:
    """追加一行数据（按 Excel 文件实际表头顺序写入）"""
    with _lock:
        wb = load_workbook(filepath)
        ws = wb.active
        file_headers = [cell.value for cell in ws[1]]
        ws.append([data.get(h, "") for h in file_headers])
        wb.save(filepath)


def _update_row(filepath: str, headers: list, key_col: str, key_val, updates: dict) -> bool:
    """按主键更新一行"""
    key_idx = headers.index(key_col) + 1
    with _lock:
        wb = load_workbook(filepath)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[key_idx - 1].value) == str(key_val):
                for col_name, val in updates.items():
                    if col_name in headers:
                        col_idx = headers.index(col_name) + 1
                        ws.cell(row=row[0].row, column=col_idx, value=val)
                wb.save(filepath)
                return True
        wb.save(filepath)
    return False


def _next_id(filepath: str, headers: list, id_col: str, prefix: str) -> str:
    """生成自增 ID，格式: PREFIX-0001"""
    rows = _read_all_rows(filepath, headers)
    max_num = 0
    for r in rows:
        val = str(r.get(id_col, ""))
        if val.startswith(prefix + "-"):
            try:
                num = int(val.split("-")[1])
                max_num = max(max_num, num)
            except (ValueError, IndexError):
                pass
    return f"{prefix}-{max_num + 1:04d}"


# --------------------------------------------------
# 库存操作
# --------------------------------------------------

def get_all_items() -> list[dict]:
    """获取全部库存商品"""
    return _read_all_rows(config.INVENTORY_FILE, INVENTORY_HEADERS)


def get_item(item_id: str) -> Optional[dict]:
    """按 ID 获取单个商品"""
    for item in get_all_items():
        if str(item.get("item_id")) == item_id:
            return item
    return None


def search_items(keyword: str) -> list[dict]:
    """模糊搜索（名称/分类/位置）"""
    keyword = keyword.lower()
    results = []
    for item in get_all_items():
        searchable = f"{item.get('name','')} {item.get('category','')} {item.get('location','')}".lower()
        if keyword in searchable:
            results.append(item)
    return results


def get_categories() -> list[str]:
    """获取所有分类"""
    cats = set()
    for item in get_all_items():
        cat = item.get("category", "")
        if cat:
            cats.add(str(cat))
    return sorted(cats)


def create_item(name: str, category: str, quantity: int = 0,
                unit: str = "个", location: str = "", min_stock: int = 0) -> dict:
    """新建商品"""
    init_data_files()
    item_id = _next_id(config.INVENTORY_FILE, INVENTORY_HEADERS, "item_id", "ITEM")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    item = {
        "item_id": item_id,
        "name": name,
        "category": category,
        "quantity": quantity,
        "unit": unit,
        "location": location,
        "min_stock": min_stock,
        "created_at": now,
        "updated_at": now,
    }
    _append_row(config.INVENTORY_FILE, INVENTORY_HEADERS, item)
    return item


def stock_in(item_id: str, quantity: int, operator: str = "系统", note: str = "") -> dict:
    """入库"""
    init_data_files()
    item = get_item(item_id)
    if not item:
        raise ValueError(f"商品 {item_id} 不存在")
    if quantity <= 0:
        raise ValueError("入库数量必须大于 0")

    new_qty = int(item.get("quantity", 0)) + quantity
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _update_row(config.INVENTORY_FILE, INVENTORY_HEADERS, "item_id", item_id,
                {"quantity": new_qty, "updated_at": now})

    tx_id = _next_id(config.TRANSACTIONS_FILE, TRANSACTION_HEADERS, "tx_id", "TX")
    tx = {
        "tx_id": tx_id,
        "item_id": item_id,
        "item_name": item.get("name", ""),
        "type": "in",
        "quantity": quantity,
        "balance_after": new_qty,
        "operator": operator,
        "recipient": "",
        "note": note,
        "timestamp": now,
    }
    _append_row(config.TRANSACTIONS_FILE, TRANSACTION_HEADERS, tx)
    return {"tx_id": tx_id, "item_id": item_id, "name": item.get("name"), "new_quantity": new_qty}


def stock_out(item_id: str, quantity: int, operator: str = "系统",
              recipient: str = "", note: str = "") -> dict:
    """出库"""
    init_data_files()
    item = get_item(item_id)
    if not item:
        raise ValueError(f"商品 {item_id} 不存在")
    if quantity <= 0:
        raise ValueError("出库数量必须大于 0")

    current_qty = int(item.get("quantity", 0))
    if quantity > current_qty:
        raise ValueError(f"库存不足：当前 {current_qty}，请求出库 {quantity}")

    new_qty = current_qty - quantity
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _update_row(config.INVENTORY_FILE, INVENTORY_HEADERS, "item_id", item_id,
                {"quantity": new_qty, "updated_at": now})

    tx_id = _next_id(config.TRANSACTIONS_FILE, TRANSACTION_HEADERS, "tx_id", "TX")
    tx = {
        "tx_id": tx_id,
        "item_id": item_id,
        "item_name": item.get("name", ""),
        "type": "out",
        "quantity": quantity,
        "balance_after": new_qty,
        "operator": operator,
        "recipient": recipient,
        "note": note,
        "timestamp": now,
    }
    _append_row(config.TRANSACTIONS_FILE, TRANSACTION_HEADERS, tx)
    return {"tx_id": tx_id, "item_id": item_id, "name": item.get("name"), "new_quantity": new_qty}


def get_transactions(item_id: str = None, tx_type: str = None,
                     limit: int = 50) -> list[dict]:
    """查询出入库记录"""
    rows = _read_all_rows(config.TRANSACTIONS_FILE, TRANSACTION_HEADERS)
    if item_id:
        rows = [r for r in rows if str(r.get("item_id")) == item_id]
    if tx_type:
        rows = [r for r in rows if r.get("type") == tx_type]
    rows.sort(key=lambda r: r.get("timestamp") or "", reverse=True)
    return rows[:limit]


def get_low_stock_items() -> list[dict]:
    """获取低于最低库存的商品"""
    items = get_all_items()
    return [
        it for it in items
        if int(it.get("min_stock", 0)) > 0
        and int(it.get("quantity", 0)) < int(it.get("min_stock", 0))
    ]
