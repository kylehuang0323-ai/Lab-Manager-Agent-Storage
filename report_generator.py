"""
报表生成器 — 生成库存清单和出入库流水的 Excel 报表
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config
import inventory_manager as im


def _style_header(ws, headers: list):
    """统一表头样式"""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = max(len(str(header)) * 2 + 4, 14)


def export_inventory_report(category: str = None) -> str:
    """导出库存清单 Excel"""
    items = im.get_all_items()
    if category:
        items = [it for it in items if str(it.get("category", "")) == category]

    wb = Workbook()
    ws = wb.active
    ws.title = "库存清单"

    headers = ["商品编号", "名称", "分类", "数量", "单位", "存放位置", "最低库存", "状态", "更新时间"]
    _style_header(ws, headers)

    warn_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    danger_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for i, item in enumerate(items, 2):
        qty = int(item.get("quantity", 0))
        min_s = int(item.get("min_stock", 0))
        if min_s > 0 and qty == 0:
            status = "⚠️ 缺货"
            row_fill = danger_fill
        elif min_s > 0 and qty < min_s:
            status = "⚠️ 低库存"
            row_fill = warn_fill
        else:
            status = "✅ 正常"
            row_fill = None

        row_data = [
            item.get("item_id"), item.get("name"), item.get("category"),
            qty, item.get("unit"), item.get("location"),
            min_s, status, item.get("updated_at"),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            if row_fill:
                cell.fill = row_fill

    # 汇总行
    summary_row = len(items) + 3
    ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=f"共 {len(items)} 种商品")
    total_qty = sum(int(it.get("quantity", 0)) for it in items)
    ws.cell(row=summary_row, column=3, value=f"总数量: {total_qty}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inventory_report_{timestamp}.xlsx"
    filepath = os.path.join(config.EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath


def export_transactions_report(item_id: str = None, tx_type: str = None,
                                limit: int = 200) -> str:
    """导出出入库流水 Excel"""
    records = im.get_transactions(item_id=item_id, tx_type=tx_type, limit=limit)

    wb = Workbook()
    ws = wb.active
    ws.title = "出入库记录"

    headers = ["流水号", "商品编号", "商品名称", "类型", "数量", "操作人", "领用人", "备注", "时间"]
    _style_header(ws, headers)

    in_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    out_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    for i, tx in enumerate(records, 2):
        tx_t = tx.get("type", "")
        row_fill = in_fill if tx_t == "in" else out_fill
        row_data = [
            tx.get("tx_id"), tx.get("item_id"), tx.get("item_name"),
            "入库" if tx_t == "in" else "出库",
            tx.get("quantity"), tx.get("operator"),
            tx.get("recipient", ""), tx.get("note", ""), tx.get("timestamp"),
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.fill = row_fill

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"transactions_report_{timestamp}.xlsx"
    filepath = os.path.join(config.EXPORT_DIR, filename)
    wb.save(filepath)
    return filepath
